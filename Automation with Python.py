import openpyxl as xl

from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # wb = workbook
    wb = xl.load_workbook(filename)
    # using square brackets means targeting inside of the subject, in this case
    # we target Sheet1 inside wb
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_value = cell.value * 0.8
        # adding the corrected_value to the 4th column
        corrected_value_cell = sheet.cell(row, 4)
        corrected_value_cell.value = corrected_value

    values_chart = Reference(sheet,
                             min_row=2,
                             max_row=sheet.max_row,
                             min_col=4,
                             max_col=4)
    chart = BarChart()
    chart.add_data(values_chart)
    # this includes the coordinates of the top left of the chart which in this case is e2
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook('transactions.xlsx')
